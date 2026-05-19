"""
Exporter — filters a DataFrame by branch and writes per-branch Excel attachments.

Each branch email gets its own filtered .xlsx file as an attachment.
Files are named after the email subject (sanitised for filesystem safety).
Files can be automatically deleted after sending (engine auto_delete setting).
"""

import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


class DispatchExporter:
    """
    Converts a filtered Pandas DataFrame into a styled Excel file.

    The exporter writes the filtered branch data to a single sheet named 'Data',
    applies a styled header row (bold, light-blue fill), auto-fits column widths,
    and freezes the header row for easy scrolling.

    File naming
    -----------
    The output filename is derived from the email subject by sanitising
    characters that are invalid in filenames (<>:"/\\|?*). The file is saved
    as subject-sanitised.xlsx in the upload_folder directory.

    Cleanup
    -------
    Call exporter.cleanup(filepath) to delete the file after sending.
    The engine handles this automatically when auto_delete=True.
    """

    def export(self, data, subject: str, upload_folder: str) -> str:
        """
        Write a branch-filtered DataFrame to an Excel file.

        Parameters
        ----------
        data : pd.DataFrame
            Filtered DataFrame for one branch (rows matching branch name).
        subject : str
            Email subject line. Used to derive the output filename after sanitisation.
        upload_folder : str
            Directory to write the file to. Must already exist.

        Returns
        -------
        str
            Absolute path to the written .xlsx file.
        """
        safe_name = self._sanitise_filename(subject)
        filename = f"{safe_name}.xlsx"
        filepath = os.path.join(upload_folder, filename)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Data"

        # ── Header row ─────────────────────────────────────────────────────
        headers = list(data.columns)
        ws.append(headers)

        # Style: bold, light-blue (#D6E4F0), centred, bordered
        header_fill = PatternFill(
            start_color="D6E4F0",
            end_color="D6E4F0",
            fill_type="solid"
        )
        header_font = Font(bold=True, color="1F2937", name="Calibri", size=10)
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin = Side(style='thin', color="BFBFBF")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for cell in ws[1]:
            cell.font      = header_font
            cell.fill      = header_fill
            cell.alignment = header_alignment
            cell.border    = border

        # ── Data rows ──────────────────────────────────────────────────────
        for row in data.itertuples(index=False, name=None):
            ws.append(row)

        # ── Auto-fit column widths ─────────────────────────────────────────
        # Cap at 40 characters to prevent extremely wide columns.
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_length + 2, 40)

        # ── Freeze header row ────────────────────────────────────────────────
        ws.freeze_panes = "A2"

        wb.save(filepath)
        return filepath

    @staticmethod
    def _sanitise_filename(name: str) -> str:
        """
        Remove or replace characters that are invalid in filenames.
        Windows and Linux both restrict: < > : " / \\ | ? * and control chars.
        Result is truncated to 100 characters.
        """
        invalid = '<>:"/\\|?*'
        for char in invalid:
            name = name.replace(char, '_')
        name = name.strip()
        return name[:100] if name else "attachment"

    @staticmethod
    def cleanup(filepath: str):
        """
        Delete a temporary attachment file if it still exists.
        Called by the engine after each email is sent (when auto_delete=True).
        Silently ignores errors (file already deleted, permissions, etc.).
        """
        try:
            if os.path.exists(filepath):
                os.remove(filepath)
        except Exception:
            pass