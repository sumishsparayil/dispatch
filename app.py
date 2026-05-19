"""
Dispatch — Flask Web Application
KLM Axiva Finvest — MIS Email Dispatch System

Handles all HTTP routes (pages + API) and orchestrates the dispatch pipeline.
Session state lives in module scope ( cleared on /api/clear ).
SMTP config is never stored in branch_mappings — address_book is the single store.
"""

import gc
import os
import sys
import json
import time as _time
import threading
import queue

# ── App directory (dev vs bundled exe) ─────────────────────────────────────
# When frozen (PyInstaller exe), sys.executable points to the .exe itself.
# When running as .py, __file__ points to app.py → use its directory.
if getattr(sys, 'frozen', False):
    APP_DIR = os.path.dirname(os.path.abspath(sys.executable))
else:
    APP_DIR = os.path.dirname(os.path.abspath(__file__))

os.chdir(APP_DIR)

from flask import Flask, render_template, request, jsonify, Response
from werkzeug.utils import secure_filename
import pandas as pd
import uuid

from core.parser import PantherParser
from core.engine import PantherEngine
from core.mailer import PantherMailer
from core.exporter import PantherExporter
from db.database import (
    init_db, get_settings, save_settings_batch, get_password_for_smtp,
    get_run_history, log_run_start, log_run_complete, log_email_result,
    get_run_detail,
)
from db.address_book import (
    get_address_book, save_address_book,
    add_branch, update_branch, delete_branch,
    import_from_excel, export_to_excel,
)

# ── Default from name ──────────────────────────────────────────────────────────
# Single source of truth. Used when the user has not set a custom from name.
DEFAULT_FROM_NAME = 'KLM Axiva MIS'

# ── Flask app ─────────────────────────────────────────────────────────────────
app = Flask(__name__, template_folder='templates', static_folder='static')
app.config['UPLOAD_FOLDER'] = os.path.join(APP_DIR, 'uploads')
app.config['MAX_CONTENT_LENGTH'] = 200 * 102 * 1024   # 200 MB file upload limit

os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

# Ensure the JSON store and address book exist before the first request is handled.
init_db()

# ── In-memory session ──────────────────────────────────────────────────────────
# Holds the parsed DataFrame and branch list for the CURRENT upload session.
# Cleared when the user clicks "Start Over" (/api/clear).
# NOT persisted — lost on app restart.
_session = {
    'data_df':      None,      # Pandas DataFrame — full parsed Excel sheet
    'branches':     [],        # List[dict] — [{branch: str, count: int}, ...]
    'columns':      [],        # List[str] — column names from the Excel file
    'branch_col':   None,      # Exact column name for 'Branch' (e.g. 'Branch' or 'BRANCH')
    'total_rows':   0,         # Total data rows in the parsed DataFrame
    'filename':     None,      # Original filename the user uploaded
    'id':           None,      # Unique session ID (UUID hex prefix, 12 chars)
}

# ── Parse progress queue ───────────────────────────────────────────────────────
# Queue used by PantherParser to emit SSE progress events during Excel parsing.
# The background parse thread writes to this queue; the SSE endpoint reads from it.
_parse_progress_queue = queue.Queue()

# ── Current parser (weak reference via list) ─────────────────────────────────
# Wrapped in a list so the closure in the parse route can mutate it.
# Set to [None] after session is cleared to allow GC of the parser object.
_current_parser = [None]


# ══════════════════════════════════════════════════════════════════════════════
# Helper functions
# ══════════════════════════════════════════════════════════════════════════════

def _smtp_settings():
    """
    Build the SMTP config dict for the mailer.
    Reads from panther_data.json (settings store). smtp_pass falls back to
    the OS keyring if not found in JSON.
    """
    s = get_settings()
    return {
        'smtp_host': s.get('smtp_host', ''),
        'smtp_port': int(s.get('smtp_port') or 587),
        'smtp_user': s.get('smtp_user', ''),
        'smtp_pass': s.get('smtp_pass') or get_password_for_smtp(),
        'from_name': s.get('from_name', DEFAULT_FROM_NAME),
        'use_tls':   s.get('use_tls', True),
    }


# ══════════════════════════════════════════════════════════════════════════════
# Page routes
# ══════════════════════════════════════════════════════════════════════════════

@app.route('/')
def index():
    """Main dashboard — file upload + progress + branch review."""
    return render_template('index.html')


@app.route('/settings')
def settings_page():
    """Settings page — SMTP config, email template, address book."""
    return render_template('settings.html')


@app.route('/history')
def history_page():
    """Run history page — past dispatch runs with sent/failed counts."""
    return render_template('history.html')


# ══════════════════════════════════════════════════════════════════════════════
# API: Upload & Parse
# ══════════════════════════════════════════════════════════════════════════════

@app.route('/api/upload', methods=['POST'])
def api_upload():
    """
    Accept an Excel file upload and start a background parse.
    Returns immediately with session_id. Progress streamed via /api/upload-progress.

    Request:  multipart/form-data with file field 'file'
    Response: {session_id: str, filename: str, total_rows: int, total_branches: int}
    Errors:   400 if no file, 415 if not .xlsx/.xlsm
    """
    if 'file' not in request.files:
        return jsonify({'error': 'No file provided'}), 400

    file = request.files['file']
    if not file.filename:
        return jsonify({'error': 'No file selected'}), 400

    ext = os.path.splitext(file.filename)[1].lower()
    if ext not in ('.xlsx', '.xlsm'):
        return jsonify({'error': 'Only .xlsx and .xlsm files are supported'}), 415

    # Save to uploads/ with a secure name (original name kept for display)
    filename = secure_filename(file.filename)
    filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    file.save(filepath)

    # Clear any stale progress from a previous parse
    while not _parse_progress_queue.empty():
        _parse_progress_queue.get_nowait()

    # Spawn background parser
    def background_parse():
        parser = PantherParser(filepath, progress_callback=_emit_parse_progress)
        _current_parser[0] = parser
        try:
            result = parser.parse()
            # Store parsed data in the in-memory session
            _session['data_df']    = result['data_df']
            _session['branches']   = result['branches']
            _session['columns']    = result['columns']
            _session['branch_col'] = result['branch_col']
            _session['total_rows'] = result['total_data_rows']
            _session['filename']   = filename
            _session['id']         = result['session_id']
            _parse_progress_queue.put({
                'stage':    'done',
                'message':  f"Ready — {result['total_branches']} branches, {result['total_data_rows']:,} rows",
                'percent':  100,
            })
        except ValueError as e:
            _parse_progress_queue.put({'stage': 'error', 'message': str(e), 'percent': 0})
        except Exception as e:
            _parse_progress_queue.put({'stage': 'error', 'message': f'Unexpected error: {e}', 'percent': 0})
        finally:
            # Always clean up the uploaded file after parsing (unless debugging)
            try:
                os.remove(filepath)
            except OSError:
                pass

    threading.Thread(target=background_parse, daemon=True).start()

    return jsonify({'filename': filename, 'message': 'Parsing started...'})


def _emit_parse_progress(stage, message, percent):
    """SSE progress callback registered with PantherParser during parse."""
    _parse_progress_queue.put({'stage': stage, 'message': message, 'percent': percent})


@app.route('/api/upload-progress')
def api_upload_progress():
    """
    SSE stream — yields parse progress events as Server-Sent Events.
    Clients should connect before calling /api/upload.
    Each event is a JSON object: {stage, message, percent}
    """
    def generate():
        while True:
            try:
                event = _parse_progress_queue.get(timeout=60)
                yield f"data: {json.dumps(event)}\n\n"
                if event.get('stage') in ('done', 'error'):
                    break
            except queue.Empty:
                yield f"data: {json.dumps({'stage': 'timeout', 'message': 'Parse timeout', 'percent': 0})}\n\n"
                break

    return Response(generate(), mimetype='text/event-stream')


@app.route('/api/session')
def api_session():
    """
    Return the current in-memory session state.
    Used by the UI to populate the branch review checklist after upload.

    Response: {
        id:           str  or null,
        filename:     str  or null,
        total_rows:   int,
        total_branches: int,
        branches:     [{branch: str, count: int}, ...],
        columns:      [str, ...],
        branch_col:   str  or null,
    }
    """
    return jsonify({
        'id':             _session['id'],
        'filename':       _session['filename'],
        'total_rows':     _session['total_rows'],
        'total_branches': len(_session['branches']),
        'branches':       _session['branches'],
        'columns':        _session['columns'],
        'branch_col':     _session['branch_col'],
    })


# ══════════════════════════════════════════════════════════════════════════════
# API: Dispatch
# ══════════════════════════════════════════════════════════════════════════════

@app.route('/api/run', methods=['POST'])
def api_run():
    """
    Start the email dispatch run for the current session.
    Must be called AFTER /api/session returns a non-null id.

    Request body (JSON, optional):
        excluded_branches: [str, ...]   — branches unchecked by user
        test_only:         bool         — send to operator's own email only

    Response: {run_id: str, total_branches: int}
    """
    if _session['id'] is None:
        return jsonify({'error': 'No file uploaded. Please upload a file first.'}), 400

    data = request.get_json() or {}
    excluded_branches = data.get('excluded_branches', [])
    test_only        = data.get('test_only', False)

    # Build full settings dict for the engine.
    # branch_mappings is built from address_book.json — the single authoritative store.
    ab             = get_address_book()
    branch_mappings = {
        str(branch): {'email': info.get('email', ''), 'cc': info.get('cc', '')}
        for branch, info in ab.get('branches', {}).items()
        if info.get('email')
    }
    settings = {**get_settings(), 'branch_mappings': branch_mappings}

    engine = PantherEngine(
        branches          = _session['branches'],
        data_df           = _session['data_df'],
        branch_col        = _session['branch_col'],
        settings          = settings,
        upload_folder     = app.config['UPLOAD_FOLDER'],
        filename          = _session['filename'] or '',
        excluded_branches = excluded_branches,
        test_only         = test_only,
    )

    # Test mode: override all recipients → send only to the operator's own email.
    # The engine's test_only flag redirects to the SMTP user (operator's address).
    if test_only:
        engine.settings['test_recipient'] = settings.get('smtp_user', '')

    engine.start()
    return jsonify({'run_id': engine.run_id, 'total_branches': len(_session['branches'])})


@app.route('/api/progress')
def api_progress():
    """
    SSE stream — yields dispatch log entries as Server-Sent Events.
    Each event is a JSON object: {event, message, level, timestamp, ...}
    """
    last_log_len = [0]

    def generate():
        while True:
            log_entries = PantherEngine.get_log()
            if len(log_entries) > last_log_len[0]:
                for entry in log_entries[last_log_len[0]:]:
                    yield f"data: {json.dumps(entry)}\n\n"
                    last_log_len[0] = len(log_entries)
                    if entry.get('done'):
                        return
            _time.sleep(0.3)

    return Response(generate(), mimetype='text/event-stream')


@app.route('/api/stop', methods=['POST'])
def api_stop():
    """
    Signal the running engine to stop gracefully.
    The engine checks _stop_flag between each branch and exits the loop.
    """
    PantherEngine.stop()
    return jsonify({'message': 'Stop signalled'})


# ══════════════════════════════════════════════════════════════════════════════
# API: Settings
# ══════════════════════════════════════════════════════════════════════════════

@app.route('/api/settings', methods=['GET'])
def api_settings_get():
    """
    Return all settings from panther_data.json.
    Omits branch_mappings (never stored here — address_book is the only store).
    """
    settings = get_settings()
    # Never leak smtp_pass to the frontend — return only stored flag
    return jsonify({k: v for k, v in settings.items() if k != 'smtp_pass'})


@app.route('/api/settings', methods=['POST'])
def api_settings_save():
    """
    Save one or more settings to panther_data.json.
    branch_mappings is silently ignored — address_book is the authoritative store.
    smtp_pass is stored directly (plain text — acceptable for internal tool behind VPN).
    """
    data = request.get_json()
    if not data:
        return jsonify({'error': 'No settings provided'}), 400

    # Strip branch_mappings silently — never save it to the settings store.
    save_settings_batch({k: v for k, v in data.items() if k != 'branch_mappings'})
    return jsonify({'message': 'Settings saved'})


@app.route('/api/test-smtp', methods=['POST'])
def api_test_smtp():
    """
    Send a minimal test email to verify SMTP connectivity.
    Sent to the operator's own SMTP address (smtp_user from settings).
    """
    data     = request.get_json() or {}
    recipient = data.get('recipient', '') or get_settings().get('smtp_user', '')
    settings = _smtp_settings()
    mailer   = PantherMailer(settings)

    result = mailer.send_test(to=recipient)
    if result.get('ok'):
        return jsonify({'message': f'Test email sent to {recipient}'})
    return jsonify({'error': result.get('error', 'Unknown error')}), 500


# ══════════════════════════════════════════════════════════════════════════════
# API: Address Book
# ══════════════════════════════════════════════════════════════════════════════

@app.route('/api/address-book', methods=['GET'])
def api_address_book_get():
    """Return the full address book (branches dict + default_cc)."""
    return jsonify(get_address_book())


@app.route('/api/address-book/branches', methods=['POST'])
def api_address_book_upsert():
    """
    Add or update a single branch entry.
    Request body: {branch: str, email: str, cc: str}
    """
    data = request.get_json()
    if not data:
        return jsonify({'error': 'Branch data required'}), 400
    update_branch(
        branch = data.get('branch', '').strip(),
        email  = data.get('email', '').strip(),
        cc     = data.get('cc', '').strip(),
    )
    return jsonify({'message': 'Branch updated'})


@app.route('/api/address-book/branches/<branch>', methods=['DELETE'])
def api_address_book_delete(branch):
    """Remove a branch from the address book."""
    delete_branch(branch)
    return jsonify({'message': f'{branch} removed'})


@app.route('/api/address-book/import', methods=['POST'])
def api_address_book_import():
    """
    Import branches from an Excel file (multipart/form-data, field 'file').
    Expected columns: Branch, Email, CC (column names are case-insensitive).
    Returns: {added: int, updated: int, errors: [str, ...]}
    """
    if 'file' not in request.files:
        return jsonify({'error': 'No file provided'}), 400

    file = request.files['file']
    ext  = os.path.splitext(file.filename)[1].lower()
    if ext not in ('.xlsx', '.xlsm'):
        return jsonify({'error': 'Only .xlsx and .xlsm files are supported'}), 415

    filepath = os.path.join(app.config['UPLOAD_FOLDER'], secure_filename(file.filename))
    file.save(filepath)

    try:
        result = import_from_excel(filepath)
        return jsonify(result)
    finally:
        try:
            os.remove(filepath)
        except OSError:
            pass


@app.route('/api/address-book/export', methods=['GET'])
def api_address_book_export():
    """Download the address book as a .xlsx file."""
    buf = export_to_excel()
    return Response(
        buf.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': 'attachment; filename=address_book.xlsx'},
    )


@app.route('/api/address-book/fill-from-session', methods=['POST'])
def api_fill_from_session():
    """
    For every branch in the current session that has NO email in the address book,
    pre-fill the email field with an empty string as a placeholder.
    Allows the operator to fill in emails directly in the address book tab.
    """
    ab = get_address_book()
    added = 0
    for b in _session.get('branches', []):
        name = b['branch']
        if name not in ab.get('branches', {}):
            add_branch(branch=name, email='', cc='')
            added += 1
    return jsonify({'message': f'{added} branch placeholders added'})


# ══════════════════════════════════════════════════════════════════════════════
# API: History
# ══════════════════════════════════════════════════════════════════════════════

@app.route('/api/history', methods=['GET'])
def api_history():
    """Return the list of past dispatch runs (most recent first)."""
    return jsonify(get_run_history())


@app.route('/api/history/<run_id>', methods=['GET'])
def api_history_detail(run_id):
    """
    Return per-email records for a specific run.
    Each record: {recipient, cc, subject, branch, status, error_message, sent_at}
    """
    return jsonify(get_run_detail(run_id))


# ══════════════════════════════════════════════════════════════════════════════
# API: Clear Session
# ══════════════════════════════════════════════════════════════════════════════

@app.route('/api/clear', methods=['POST'])
def api_clear():
    """
    Clear the in-memory session and release RAM.
    Called when the user clicks "Start Over".
    Idempotent — safe to call even if no session is active.
    """
    _session['data_df']    = None
    _session['branches']   = []
    _session['columns']    = []
    _session['branch_col'] = None
    _session['total_rows'] = 0
    _session['filename']   = None
    _session['id']        = None

    # Release the parser reference so it can be garbage collected.
    _current_parser[0] = None
    PantherEngine.clear_log()
    gc.collect()

    return jsonify({'message': 'Session cleared'})


# ══════════════════════════════════════════════════════════════════════════════
# API: Download Reference Template
# ══════════════════════════════════════════════════════════════════════════════

@app.route('/api/download-template', methods=['GET'])
def api_download_template():
    """
    Generate and serve a blank reference Excel template showing the expected
    column structure for the data file (Branch column highlighted).
    """
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = 'Data'

    headers = ['Branch', 'LOAN AMOUNT', 'PRINCIPAL OUTSTANDING', 'DUE DATE']
    ws.append(headers)

    # Style the header row so it matches what the parser expects.
    header_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
    header_font = Font(bold=True, color='1F2937', name='Calibri', size=10)
    header_align = Alignment(horizontal='center', vertical='center')

    for cell in ws[1]:
        cell.font  = header_font
        cell.fill  = header_fill
        cell.alignment = header_align

    # Add one example row so the user knows the expected data format.
    ws.append(['PL-ALAPPUZHA', 0, 0, '2026-03-31'])
    ws.append(['PL-ERNAKULAM', 0, 0, '2026-03-31'])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return Response(
        buf.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': 'attachment; filename=dispatch_template.xlsx'},
    )


# ══════════════════════════════════════════════════════════════════════════════
# Run
# ══════════════════════════════════════════════════════════════════════════════

if __name__ == '__main__':
    import argparse

    parser = argparse.ArgumentParser(description='Dispatch MIS Email System')
    parser.add_argument('--port', type=int, default=5000, help='Port to run on (default: 5000)')
    parser.add_argument('--host', default='0.0.0.0', help='Host to bind to (default: 0.0.0.0)')
    args = parser.parse_args()

    print(f"Starting Dispatch on {args.host}:{args.port}")
    app.run(host=args.host, port=args.port, debug=False, threaded=True)